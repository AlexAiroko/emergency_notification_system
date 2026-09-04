import asyncio
from io import BytesIO
import logging

from fastapi import UploadFile
from openpyxl import load_workbook

from app.exceptions.contact_import import EmptyImportFileError
from app.services.contact_import.parsers.base import BaseContactParser


logger = logging.getLogger(__name__)


class ExcelParser(BaseContactParser):
    async def parse(self, file: UploadFile) -> list[dict]:
        logger.debug("Parsing Excel file: %s", file.filename)
        await file.seek(0)
        content = await file.read()

        workbook = await asyncio.to_thread(
            load_workbook, BytesIO(content)
        )
        
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            raise EmptyImportFileError()

        headers = list(rows[0])
        
        self.validate_headers(headers)
        
        return [dict(zip(headers, values)) for values in rows[1:]]
